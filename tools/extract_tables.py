# -*- coding: utf-8 -*-
"""Excel'den sabit veri tablolarını JSON'a çıkarır.

Kaynak: 11.Tayakadın Deresi SENTETİK YÖNTEMLER TABLOLU.xlsm
Çıktı:  data/tables/*.json
"""
import json
import os
import re
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSM = os.path.join(os.path.dirname(ROOT), "11.Tayakadın Deresi SENTETİK YÖNTEMLER TABLOLU.xlsm")
OUT = os.path.join(ROOT, "data", "tables")
os.makedirs(OUT, exist_ok=True)


def dump(name, obj):
    path = os.path.join(OUT, name + ".json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=1)
    print("yazildi:", path)


def main():
    wbf = openpyxl.load_workbook(XLSM, data_only=False)  # formüller
    wbv = openpyxl.load_workbook(XLSM, data_only=True)   # değerler

    # ---- 1) SCS/DSİ boyutsuz birim hidrograf (BH2 Hesabı A5:B..) ----
    ws = wbv["BH2 Hesabı"]
    t_tp, q_qp = [], []
    r = 5
    while ws.cell(r, 1).value is not None:
        t_tp.append(float(ws.cell(r, 1).value))
        q_qp.append(float(ws.cell(r, 2).value))
        r += 1
    dump("bh2_dimensionless", {"t_tp": t_tp, "q_qp": q_qp})

    # ---- 2) YZD bölge eğrileri (DATAGİR artım akış formüllerinden) ----
    wsf = wbf["DATAGİR"]
    # grup: (etiket kolonu, yzd kolonu)
    groups = [("A", "C"), ("J", "L"), ("S", "U"), ("AB", "AD"), ("AK", "AM"), ("AT", "AV")]
    pat_label = re.compile(r"^(\d+)/(\d+)$")
    pat_yzd = re.compile(
        r'IF\(\$L\$4="A",([\d.]+),IF\(\$L\$4="B",([\d.]+),IF\(\$L\$4="C",([\d.]+)'
    )
    curve = {}  # ratio -> {A,B,C}
    for row in range(60, 200):
        for lc, yc in groups:
            lab = wsf[f"{lc}{row}"].value
            if not isinstance(lab, str):
                continue
            m = pat_label.match(lab.strip())
            if not m:
                continue
            num, den = int(m.group(1)), int(m.group(2))
            ratio = round(num / den, 6)
            yv = wsf[f"{yc}{row}"].value
            if isinstance(yv, (int, float)):
                curve[ratio] = {"A": float(yv), "B": float(yv), "C": float(yv)}
            elif isinstance(yv, str):
                mm = pat_yzd.search(yv.replace(" ", ""))
                if mm:
                    curve[ratio] = {
                        "A": float(mm.group(1)),
                        "B": float(mm.group(2)),
                        "C": float(mm.group(3)),
                    }
    ratios = sorted(curve)
    dump("yzd_curves", {
        "aciklama": "24 saatten kısa süreli yağış oranları; t/D oranına göre bölge (A/B/C) katsayısı",
        "ratios": ratios,
        "A": [curve[r]["A"] for r in ratios],
        "B": [curve[r]["B"] for r in ratios],
        "C": [curve[r]["C"] for r in ratios],
    })

    # ---- 3) ABAK2: alansal azaltma (YAD) ----
    ws = wbv["ABAK2"]
    durations = [float(ws.cell(2, c).value) for c in range(2, 7)]  # B2:F2 (saat)
    areas, vals = [], []
    r = 3
    while ws.cell(r, 1).value is not None:
        row_vals = [ws.cell(r, c).value for c in range(2, 7)]
        if all(isinstance(v, (int, float)) for v in row_vals):
            areas.append(float(ws.cell(r, 1).value))
            vals.append([float(v) for v in row_vals])
        r += 1
    dump("abak2_yad", {"durations_hr": durations, "areas_km2": areas, "percent": vals})

    # ---- 4) DPLV istasyon dağılım eğrileri (DATAGİR P23:AD62) ----
    ws = wbv["DATAGİR"]
    durations_min = [5, 10, 15, 30, 60, 120, 180, 240, 300, 360, 480, 720, 1080, 1440]
    stations = []
    for r in range(23, 63):
        name = ws.cell(r, 16).value  # P
        vals = [ws.cell(r, c).value for c in range(17, 31)]  # Q..AD
        if name and all(isinstance(v, (int, float)) for v in vals):
            stations.append({"name": str(name).strip(), "ratios": [float(v) for v in vals]})
    dump("dplv_stations", {"durations_min": durations_min, "stations": stations})

    # ---- 5) CN Şart II -> Şart III dönüşümü ----
    ws = wbv["CNIII"]
    cn2, cn3 = [], []
    r = 3
    while ws.cell(r, 1).value is not None:
        cn2.append(float(ws.cell(r, 1).value))
        cn3.append(float(ws.cell(r, 2).value))
        r += 1
    dump("cn2_to_cn3", {"cn2": cn2, "cn3": cn3})

    # ---- 6) Golden test verisi (Tayakadın) ----
    wsd = wbv["DATAGİR"]
    golden = {
        "girdi": {
            "ad": "TAYAKADIN DERESİ",
            "A_km2": wsd["A6"].value,
            "L_km": wsd["B6"].value,
            "Lc_km": wsd["C6"].value,
            "CN2": wsd["D6"].value,
            "CN3": wsd["E6"].value,
            "bolge": wsd["L4"].value,
            "kotlar": [wsd.cell(14, c).value for c in range(1, 12)],  # A14:K14
            "Qbaz": wsd["G10"].value,
            "dplv_no": wsd["O21"].value,
            "P24_agirlikli": [wsd.cell(22, c).value for c in range(1, 8)],  # A22:G22 (2..100 + OEY)
            "istasyon_P24": {
                str(wsd.cell(r, 9).value).strip(): [wsd.cell(r, c).value for c in range(1, 8)]
                for r in range(23, 28) if wsd.cell(r, 9).value
            },
            "thiessen_agirlik_H26": wsd["H26"].value,
        },
        "beklenen": {},
    }
    wsk = wbv["KABULET"]
    dur = [2, 4, 6, 8, 12, 18, 24]
    mat = {}
    for i, rp in enumerate(["Q2", "Q5", "Q10", "Q25", "Q50", "Q100", "QOET"]):
        mat[rp] = [wsk.cell(7 + i, 2 + j).value for j in range(7)]
    golden["beklenen"]["kabulet_pik"] = {"sureler_saat": dur, "matris": mat}
    wso = wbv["Önhesap"]
    golden["beklenen"]["dsi"] = {
        "S_harmonik": wso["D6"].value,
        "qp_l_s_km2_mm": wso["D25"].value,
        "Qp_m3_s_mm": wso["C28"].value,
        "T_saat": wso["E32"].value,
        "Tp_saat": wso["C34"].value,
    }
    wsm = wbv["MOCKUS"]
    golden["beklenen"]["mockus"] = {
        "Tc": wsm["B4"].value,
        "D": wsm["D5"].value,
        "Tp": wsm["J5"].value,
        "K3": wsm["L6"].value,
        "qp_K1": wsm["H7"].value,
        "qp_K2": wsm["J7"].value,
        "qp_K3": wsm["L7"].value,
    }
    wsb = wbv["BH2 Hesabı"]
    uh_t, uh_q = [], []
    for r in range(5, 120):
        tv, qv = wsb.cell(r, 19).value, wsb.cell(r, 21).value  # S, U
        if tv is None:
            break
        uh_t.append(tv)
        uh_q.append(qv)
    golden["beklenen"]["dsi_birim_hidrograf"] = {"t": uh_t, "q": uh_q}
    dump("golden_tayakadin", golden)

    print("TAMAM")


if __name__ == "__main__":
    sys.exit(main())

# -*- coding: utf-8 -*-
"""MGM PLV 2020 son2.xlsx'ten istasyon yağış (P24) ve PLV oranlarını çıkarır.

Her numaralı sayfada:
  - süre başlığı (satır: [None,5,10,15,30,1,2,3,4,5,6,8,12,18,24,*]) — 24 sa = kolon 15
  - "2 YIL".."500 YIL" satırları: her süre için tekerrürlü yağış (mm); 24 sa = kolon 15
  - "PLV" satırı: 14 oran (5,10,15,30 dk + 1,2,3,4,5,6,8,12,18,24 sa) = kolon 2..15
Çıktı: data/tables/mgm_plv_2020.json
"""
import json
import os
import re

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "MGM PLV 2020 son2.xlsx")
OUT = os.path.join(ROOT, "data", "tables", "mgm_plv_2020.json")

RP_LABELS = {"2 YIL": "2", "5 YIL": "5", "10 YIL": "10", "25 YIL": "25",
             "50 YIL": "50", "100 YIL": "100", "200 YIL": "200", "500 YIL": "500"}
COL_24H = 15  # 24 saatlik kolon (1-indeksli)


def clean_name(raw):
    """'ACIPAYAM 2020 PLV' -> 'ACIPAYAM'."""
    return re.sub(r"\s*2020\s*(PLV|PLF)\s*$", "", str(raw).strip(), flags=re.I).strip()


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    # PLV özet sayfasından S# -> ad
    plv = wb["PLV"]
    names = {}
    for r in range(3, plv.max_row + 1):
        s = plv.cell(r, 1).value
        nm = plv.cell(r, 2).value
        if isinstance(s, (int, float)) and nm:
            names[int(s)] = clean_name(nm)

    stations = []
    for s in sorted(names):
        sheet = str(s)
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        p24, plv14 = {}, None
        for r in range(1, ws.max_row + 1):
            lab = ws.cell(r, 1).value
            if not isinstance(lab, str):
                continue
            lab = lab.strip()
            if lab in RP_LABELS:
                v = ws.cell(r, COL_24H).value
                if isinstance(v, (int, float)):
                    p24[RP_LABELS[lab]] = round(float(v), 2)
            elif lab.upper() == "PLV":
                vals = [ws.cell(r, c).value for c in range(2, 16)]  # 14 değer
                if all(isinstance(v, (int, float)) for v in vals):
                    plv14 = [round(float(v), 4) for v in vals]
        if len(p24) >= 6 and plv14:
            stations.append({"no": s, "ad": names[s], "P24": p24, "plv": plv14})

    stations.sort(key=lambda x: x["ad"])
    data = {
        "aciklama": "MGM 2020 PLV — istasyon 24 saatlik tekerrürlü yağışlar (mm) ve "
                    "plüviyograf (PLV) oranları (5,10,15,30 dk + 1,2,3,4,5,6,8,12,18,24 sa)",
        "sure_dk": [5, 10, 15, 30, 60, 120, 180, 240, 300, 360, 480, 720, 1080, 1440],
        "istasyonlar": stations,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)
    print(f"yazildi: {OUT} | {len(stations)} istasyon")
    print("örnek:", stations[0]["ad"], stations[0]["P24"], "plv[0:4]=", stations[0]["plv"][:4])


if __name__ == "__main__":
    main()
